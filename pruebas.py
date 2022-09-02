import openpyxl 
  
wb = openpyxl.Workbook() 
  
sheet = wb.active 
  
sheet.cell(row = 1, column = 1).value = ' hello '
  
sheet.cell(row = 2, column = 2).value = ' everyone '
  
sheet.row_dimensions[1].height = 70
  
sheet.column_dimensions['B'].width = 20
  
wb.save('dimension.xlsx') 