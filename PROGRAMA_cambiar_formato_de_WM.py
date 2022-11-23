import json
import pandas as pd
from tkinter import filedialog,messagebox
import openpyxl
from openpyxl.styles import Alignment,Font,PatternFill
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
import xlsxwriter
import os

def exportar_excel(a):
	df = pd.read_excel(f"{a}")
	#a = 'port'
	diccionario=dict()

	for columna in df.columns:
		for index, row in df.iterrows():
			diccionario[columna] = f"{row[columna]}"

	archivo = xlsxwriter.Workbook(f"{a}")
	hoja = archivo.add_worksheet()
	row=0;column=0

	for values,items in diccionario.items():

		# AÑADIR COLUMNA #
		hoja.write(row,column,values)

		# AÑADIR VALOR#
		try:
			hoja.write(row,column+1,items)
		except Exception as e:
			print(e)
			hoja.write(row,column+1,"(null)")

		# AÑADIR ESPACIO #
		hoja.write(row,column+2,"")

		if column < 10:
			column += 3
		else:
			row+=1
			column=0
	archivo.close()
	
	#Darle formato al excel
	wb = openpyxl.load_workbook(f"{a}") 
	sheet = wb.active 
	#Tamaño a las columnas
	sheet.column_dimensions['A'].width = 20
	sheet.column_dimensions['B'].width = 20
	sheet.column_dimensions['C'].width = 5
	sheet.column_dimensions['D'].width = 20
	sheet.column_dimensions['E'].width = 20
	sheet.column_dimensions['F'].width = 5
	sheet.column_dimensions['G'].width = 20
	sheet.column_dimensions['H'].width = 20
	sheet.column_dimensions['I'].width = 5
	sheet.column_dimensions['J'].width = 20
	sheet.column_dimensions['K'].width = 30
	sheet.column_dimensions['L'].width = 5
	sheet.column_dimensions['M'].width = 20
	sheet.column_dimensions['N'].width = 50
	for row in sheet:
	  for cell in row:
	    cell.alignment = Alignment(wrapText=True,horizontal='center')	#Ajusta el texto y lo centra
	#1-17, A, D, G, J, M
	lista_celdas = ['A1','A2','A3','A4','A5','A6','A7','A8','A9','A10','A11','A12','A13','A14','A15','A16','A17',
					'D1','D2','D3','D4','D5','D6','D7','D8','D9','D10','D11','D12','D13','D14','D15','D16','D17',
					'G1','G2','G3','G4','G5','G6','G7','G8','G9','G10','G11','G12','G13','G14','G15','G16','G17',
					'J1','J2','J3','J4','J5','J6','J7','J8','J9','J10','J11','J12','J13','J14','J15','J16','J17',
					'M1','M2','M3','M4','M5','M6','M7','M8','M9','M10','M11','M12','M13','M14','M15','M16','M17']
	for celda_lista in lista_celdas:
		sheet[celda_lista].font = Font(bold=True)
	redFill = PatternFill(start_color='EE1111', end_color='EE1111',fill_type='solid')
	#sheet.conditional_formatting.add('D2:D41',CellIsRule(operator='lessThan', formula=['3.0'], stopIfTrue=True, fill=redFill))
	wb.save(f"{a}") 
	
	messagebox.showinfo("Completado","Excel modificado.")


ruta_programa = os.getcwd()
try:
	a = filedialog.askopenfilename(title="Guardar", initialdir = ruta_programa,filetypes = [("Archivo excel","*.xlsx")])
	exportar_excel(a)
except Exception as e:
	print(e)